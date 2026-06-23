import io
import csv
from typing import List
from models.grupo import Grupo
from utils.calculos import Calculadora
from utils.algoritmo_divisao import AlgoritmoDivisao, Transferencia
from utils.formatacao import Formatador


class Exportador:
    def __init__(self, grupo: Grupo):
        self.grupo = grupo
        self.calc = Calculadora(grupo)
        self.divisao = AlgoritmoDivisao(grupo)
        self.fmt = Formatador()

    def gerar_csv(self) -> str:
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["Descrição", "Categoria", "Valor", "Pagador", "Participantes", "Data"])
        for d in self.grupo.despesas:
            pagador = self.grupo.get_participante(d.pagador_id)
            participantes = [
                self.grupo.get_participante(pid).nome
                for pid in d.participantes_ids
                if self.grupo.get_participante(pid)
            ]
            writer.writerow([
                d.descricao,
                d.categoria,
                f"{d.valor:.2f}",
                pagador.nome if pagador else "-",
                ", ".join(participantes),
                d.data,
            ])

        writer.writerow([])
        writer.writerow(["--- SALDOS ---"])
        writer.writerow(["Participante", "Pago", "Consumido", "Saldo"])
        for p in self.grupo.participantes:
            pago = self.calc.total_pago_por(p.id)
            consumido = self.calc.total_consumido_por(p.id)
            saldo = self.calc.saldo_participante(p.id)
            writer.writerow([p.nome, f"{pago:.2f}", f"{consumido:.2f}", f"{saldo:.2f}"])

        writer.writerow([])
        writer.writerow(["--- TRANSFERÊNCIAS ---"])
        writer.writerow(["De", "Para", "Valor"])
        transferencias = self.divisao.calcular_transferencias()
        for t in transferencias:
            writer.writerow([t.de.nome, t.para.nome, f"{t.valor:.2f}"])

        return output.getvalue()

    def gerar_excel_bytes(self) -> bytes:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = openpyxl.Workbook()

            ws = wb.active
            ws.title = "Despesas"
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="006C49", end_color="006C49", fill_type="solid")

            headers = ["Descrição", "Categoria", "Valor (R$)", "Pagador", "Participantes", "Data"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            for row_idx, d in enumerate(self.grupo.despesas, 2):
                pagador = self.grupo.get_participante(d.pagador_id)
                participantes = [
                    self.grupo.get_participante(pid).nome
                    for pid in d.participantes_ids
                    if self.grupo.get_participante(pid)
                ]
                ws.cell(row=row_idx, column=1, value=d.descricao)
                ws.cell(row=row_idx, column=2, value=d.categoria)
                ws.cell(row=row_idx, column=3, value=d.valor)
                ws.cell(row=row_idx, column=4, value=pagador.nome if pagador else "-")
                ws.cell(row=row_idx, column=5, value=", ".join(participantes))
                ws.cell(row=row_idx, column=6, value=d.data)

            ws2 = wb.create_sheet("Saldos")
            headers2 = ["Participante", "Total Pago", "Total Consumido", "Saldo"]
            for col, h in enumerate(headers2, 1):
                cell = ws2.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill

            for row_idx, p in enumerate(self.grupo.participantes, 2):
                ws2.cell(row=row_idx, column=1, value=p.nome)
                ws2.cell(row=row_idx, column=2, value=self.calc.total_pago_por(p.id))
                ws2.cell(row=row_idx, column=3, value=self.calc.total_consumido_por(p.id))
                ws2.cell(row=row_idx, column=4, value=self.calc.saldo_participante(p.id))

            ws3 = wb.create_sheet("Transferências")
            headers3 = ["De", "Para", "Valor (R$)", "Chave PIX"]
            for col, h in enumerate(headers3, 1):
                cell = ws3.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill

            transferencias = self.divisao.calcular_transferencias()
            for row_idx, t in enumerate(transferencias, 2):
                ws3.cell(row=row_idx, column=1, value=t.de.nome)
                ws3.cell(row=row_idx, column=2, value=t.para.nome)
                ws3.cell(row=row_idx, column=3, value=t.valor)
                ws3.cell(row=row_idx, column=4, value=t.para.chave_pix)

            buffer = io.BytesIO()
            wb.save(buffer)
            return buffer.getvalue()

        except ImportError:
            return b""

    def gerar_pdf_bytes(self) -> bytes:
        try:
            from fpdf import FPDF

            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Helvetica", "B", 18)
            pdf.set_text_color(0, 108, 73)
            pdf.cell(0, 12, f"DivideAi - {self.grupo.nome}", new_x="LMARGIN", new_y="NEXT")
            pdf.set_font("Helvetica", "", 10)
            pdf.set_text_color(60, 60, 60)
            pdf.cell(0, 6, f"Tipo: {self.grupo.tipo_evento} | Participantes: {len(self.grupo.participantes)}", new_x="LMARGIN", new_y="NEXT")
            pdf.ln(8)

            kpis = self.calc.kpis()
            pdf.set_font("Helvetica", "B", 12)
            pdf.set_text_color(0, 0, 0)
            pdf.cell(0, 8, "Resumo Financeiro", new_x="LMARGIN", new_y="NEXT")
            pdf.set_font("Helvetica", "", 10)
            pdf.cell(0, 6, f"Total Gasto: {self.fmt.moeda(kpis['total_gasto'])}", new_x="LMARGIN", new_y="NEXT")
            pdf.cell(0, 6, f"Media por Pessoa: {self.fmt.moeda(kpis['media_por_pessoa'])}", new_x="LMARGIN", new_y="NEXT")
            pdf.cell(0, 6, f"Maior Despesa: {kpis['maior_despesa_desc']} ({self.fmt.moeda(kpis['maior_despesa'])})", new_x="LMARGIN", new_y="NEXT")
            pdf.cell(0, 6, f"Quem Mais Pagou: {kpis['quem_mais_pagou']} ({self.fmt.moeda(kpis['valor_mais_pagou'])})", new_x="LMARGIN", new_y="NEXT")
            pdf.ln(8)

            pdf.set_font("Helvetica", "B", 12)
            pdf.cell(0, 8, "Despesas", new_x="LMARGIN", new_y="NEXT")
            pdf.set_font("Helvetica", "", 9)
            for d in self.grupo.despesas:
                pagador = self.grupo.get_participante(d.pagador_id)
                pdf.cell(0, 5, f"  {d.descricao} | {d.categoria} | {self.fmt.moeda(d.valor)} | Pago por {pagador.nome if pagador else '-'} | {d.data}", new_x="LMARGIN", new_y="NEXT")
            pdf.ln(8)

            pdf.set_font("Helvetica", "B", 12)
            pdf.cell(0, 8, "Saldos", new_x="LMARGIN", new_y="NEXT")
            pdf.set_font("Helvetica", "", 10)
            for p in self.grupo.participantes:
                saldo = self.calc.saldo_participante(p.id)
                sinal = "+" if saldo >= 0 else ""
                pdf.cell(0, 6, f"  {p.nome}: {sinal}{self.fmt.moeda(saldo)}", new_x="LMARGIN", new_y="NEXT")
            pdf.ln(8)

            pdf.set_font("Helvetica", "B", 12)
            pdf.cell(0, 8, "Transferencias Necessarias", new_x="LMARGIN", new_y="NEXT")
            pdf.set_font("Helvetica", "", 10)
            transferencias = self.divisao.calcular_transferencias()
            for t in transferencias:
                pdf.cell(0, 6, f"  {t.de.nome} -> {t.para.nome}: {self.fmt.moeda(t.valor)} (PIX: {t.para.chave_pix_formatada})", new_x="LMARGIN", new_y="NEXT")

            buffer = io.BytesIO()
            pdf.output(buffer)
            return buffer.getvalue()

        except ImportError:
            return b""
